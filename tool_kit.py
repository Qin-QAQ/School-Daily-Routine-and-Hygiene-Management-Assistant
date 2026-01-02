import base64
import warnings
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
from SQLite_funcs import *
from display_gui import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np


def remove_empty_columns(path):
    """
    删除Excel文件中第一行和第二行都为空的列（空白列），
    并将结果保存回原文件。

    参数:
        path (str): Excel文件路径（.xlsx）
    """
    # 读取Excel，保留原始索引，不把第一行当作列名
    df = pd.read_excel(path, header=None, dtype=str)

    # 确保至少有两行，否则无法判断
    if df.shape[0] < 2:
        print("警告：Excel文件少于两行，无法判断空白列。")
        return

    # 将空字符串转为 NaN，便于统一判断
    df = df.replace(r'^\s*$', np.nan, regex=True).infer_objects(copy=False)

    # 获取所有列索引
    cols_to_drop = []
    for col in df.columns:
        # 检查第0行和第1行是否都为空（NaN）
        val1 = df.iloc[0, col]
        val2 = df.iloc[1, col]
        if pd.isna(val1) and pd.isna(val2):
            cols_to_drop.append(col)

    # 删除空白列
    df_cleaned = df.drop(columns=cols_to_drop).reset_index(drop=True)

    # 保存回原文件（覆盖）
    df_cleaned.to_excel(path, index=False, header=False)

    print(f"已删除 {len(cols_to_drop)} 个空白列，文件已更新：{path}")
def decrypt_all(d: dict[str, str]) -> dict[str, str]:
    """
    这个函数用于解密字典中所有的加密文本
    :param d: 将要解密的字典
    :return: 解密后的字典
    """
    result = {}
    encryptor = FixedIVEncryptor()
    for k in d:
        key = encryptor.decrypt(k)
        value = encryptor.decrypt(d[k])
        result[key] = value
    return result


def get_value(sheet, x: int, y: int) -> str:
    return sheet[f"{get_column_letter(x)}{y}"].value


def get_users_and_passwords() -> dict[str, dict[str, str]]:
    """
    读取Excel文件中的用户名和其对应的密码
    :return: 封装好的全对应表 (该字典是加密后的) {user_level: {name: password, ...}, ...}
    """

    def sheet_to_dict(path: str) -> dict[str, str] | None:
        """
        自动解析Excel文件提取姓名-密码映射字典
        :param path: Excel文件路径
        :return: {姓名: 密码}的字典，无有效数据时返回空字典
        """
        # 初始化返回字典
        name_pwd_dict: dict[str, str] = {}

        # 加载Excel文件
        try:
            wb = load_workbook(path, data_only=True)
            sheet = wb.active
        except FileNotFoundError:
            warnings.warn(f"Excel文件不存在：{path}")
            return name_pwd_dict
        except Exception as e:
            warnings.warn(f"加载Excel文件失败：{str(e)}")
            return name_pwd_dict

        try:
            # 读取前两行数据（核心：第一行姓名，第二行密码，同列配对）
            max_col = sheet.max_column  # 获取最大列数
            for col in range(1, max_col + 1):
                # 读取当前列的姓名（第一行）和密码（第二行）
                name_cell = sheet.cell(row=1, column=col)
                pwd_cell = sheet.cell(row=2, column=col)
                name_val = name_cell.value
                pwd_val = pwd_cell.value

                # 处理非空数据
                if name_val is not None:
                    name_str = str(name_val).strip()
                    if name_str:  # 姓名非空才处理
                        pwd_str = str(pwd_val).strip() if pwd_val is not None else ""
                        name_pwd_dict[name_str] = pwd_str

            #   若前两行无数据，尝试读取单列多行（A1=姓名，A2=密码，A3=姓名2，A4=密码2...）
            if not name_pwd_dict:
                max_row = sheet.max_row
                row = 1
                while row <= max_row:
                    name_cell = sheet.cell(row=row, column=1)
                    pwd_cell = sheet.cell(row=row + 1, column=1) if row + 1 <= max_row else None
                    name_val = name_cell.value
                    pwd_val = pwd_cell.value if pwd_cell is not None else None

                    if name_val is not None:
                        name_str = str(name_val).strip()
                        if name_str:
                            pwd_str = str(pwd_val).strip() if pwd_val is not None else ""
                            name_pwd_dict[name_str] = pwd_str

                    row += 2  # 步长2，按行配对姓名和密码
                return name_pwd_dict
        except Exception as e:
            warnings.warn(f"解析Excel数据失败：{e}")
        finally:
            wb.close()  # 确保关闭工作簿

    # 封装
    user_levels: list[str] = ["Admin", "Grader", "Teacher"]
    names_and_passwords = {}
    for user_level in user_levels:
        sheets: dict[str, str] = sheet_to_dict(f"./data/{user_level}_names_and_passwords.xlsx")

        temp_names_and_passwords = {}
        for key in sheets:
            index_name = key
            temp_names_and_passwords[index_name] = sheets[index_name]
        names_and_passwords[user_level] = temp_names_and_passwords
    remove_empty_columns("data/Admin_names_and_passwords.xlsx")
    remove_empty_columns("data/Grader_names_and_passwords.xlsx")
    remove_empty_columns("data/Teacher_names_and_passwords.xlsx")

    return names_and_passwords


def get_time():
    # date to dict{"year":2025,"month": 12, "day": 13, "hour": "15", ...}
    time = datetime.datetime.now().strftime('{"y":"%Y", "m":"%m", "d":"%d", "h":"%H", "M":"%M", "s":"%S"}')
    return str_to_dict(time)


class FixedIVEncryptor:
    """支持中文/英文/数字，且相同明文+相同密钥输出相同密文的AES加密器"""

    def __init__(self,
                 key: str = """public static void main(String []args){ System.out.print("Hi python,I'm java.")}""",
                 fixed_iv: str = "public_pilot_oil"):
        """
        初始化加密器（固定IV保证相同明文输出相同密文）
        :param key: 加密密钥（任意长度，内部自动处理为32位（AES-256））
        :param fixed_iv: 固定初始向量（必须16位字符，默认值可自定义）
        """
        self.key = key.encode('utf-8')[:32].ljust(32, b'\0')
        if len(fixed_iv) != 16:
            raise ValueError("固定IV必须是16位字符（中文/英文/数字均可）")
        self.iv = fixed_iv.encode('utf-8')[:16].ljust(16, b'\0')

    def encrypt(self, plaintext: str) -> str:
        """
        加密：相同明文+相同密钥 → 相同密文（支持中/英/数字混合）
        :param plaintext: 明文（待加密字符串）
        :return: Base64编码的密文字符串（无随机成分）
        """
        try:
            plain_bytes = plaintext.encode('utf-8')
            padded_plain = pad(plain_bytes, AES.block_size, style='pkcs7')

            cipher = AES.new(self.key, AES.MODE_CBC, self.iv)
            cipher_bytes = cipher.encrypt(padded_plain)

            encrypted = base64.b64encode(cipher_bytes).decode('utf-8')
            return encrypted
        except Exception as e:
            raise ValueError(f"加密失败：{str(e)}")

    def decrypt(self, ciphertext: str) -> str:
        """
        解密：相同密文+相同密钥 → 相同明文
        :param ciphertext: 加密返回的Base64密文字符串
        :return: 还原后的明文（中/英/数字混合）
        """
        try:
            # Base64解码 → AES-CBC解密 → 去除补位 → 转回UTF-8字符串
            cipher_bytes = base64.b64decode(ciphertext)
            cipher = AES.new(self.key, AES.MODE_CBC, self.iv)
            decrypted_padded = cipher.decrypt(cipher_bytes)
            decrypted_bytes = unpad(decrypted_padded, AES.block_size, style='pkcs7')
            plaintext = decrypted_bytes.decode('utf-8')
            return plaintext
        except Exception as e:
            raise ValueError(f"解密失败：{str(e)}（请检查密钥/IV/密文是否正确）")


class User:
    def __init__(self, login_name, password):
        """

        User类用于生成一个用户对象
        :param login_name: 登录的姓名，也就是用户传入的账户
        :param password: 登录的密码
        """
        self.login_name = login_name
        self.password = password
        self.__level = "has no login"
        self.__punch_state = False

    def __get_input(self, text: str, parent_window) -> str:
        dialog = tk.Toplevel(parent_window)
        dialog.title(text)
        dialog.geometry("300x120")
        dialog.resizable(False, False)
        dialog.transient(parent_window)
        dialog.grab_set()

        # 居中
        dialog.update_idletasks()
        x = parent_window.winfo_rootx() + (parent_window.winfo_width() // 2) - 150
        y = parent_window.winfo_rooty() + (parent_window.winfo_height() // 2) - 60
        dialog.geometry(f"+{x}+{y}")

        # 手动创建 widget 并包装
        label_widget = tk.Label(dialog, text=text, anchor='w')
        field_widget = tk.Entry(dialog)
        button_widget = tk.Button(dialog, text="确定")

        from display_gui import Component

        class _WrappedLabel(Component):
            def __init__(self, w):
                super().__init__(w)

        class _WrappedField(Component):
            def __init__(self, w):
                super().__init__(w)

            def getText(self):
                return self.widget.get()

        class _WrappedButton(Component):
            def __init__(self, w):
                super().__init__(w)

            def addActionListener(self, f):
                self.widget.config(command=f)

        label = _WrappedLabel(label_widget)
        field = _WrappedField(field_widget)
        button = _WrappedButton(button_widget)

        label.setBounds(20, 20, 260, 20)
        field.setBounds(20, 50, 260, 25)
        button.setBounds(110, 85, 80, 25)

        result = [None]

        def on_submit():
            result[0] = field.getText()
            dialog.destroy()

        button.addActionListener(on_submit)
        field_widget.bind("<Return>", lambda e: on_submit())

        # ✅ 关键修复：不要设置 protocol！让默认行为处理 × 按钮
        # dialog.protocol("WM_DELETE_WINDOW", ...)  ← 删除这行！

        parent_window.wait_window(dialog)  # 阻塞直到 dialog 被 destroy

        return result[0] if result[0] is not None else ""

    @property
    def login_level(self) -> str:
        """
        User 的参数，同时刷新User的login_level
        :return: str
        """

        user_levels: list[str] = ["Admin", "Grader", "Teacher"]
        encryptor = FixedIVEncryptor()  # 加密器

        for user_level in user_levels:
            """
                index_levels 是一个列表，不是变量或字典！！！
            """
            index_level: dict[str, str] = get_users_and_passwords()[
                user_level]  # 他返回的是字典，格式：{user_name: password, user_name: password, ...}
            encrypted_name = encryptor.encrypt(self.login_name)
            encrypted_password = encryptor.encrypt(self.password)
            if encrypted_name in index_level:
                if encrypted_password == index_level[encrypted_name]:
                    temp = user_level
                    self.__level = temp
        return self.__level

    @login_level.setter
    def login_level(self, level: str):
        self.__level = level

    def add_admin(self, root):
        """
        用于管理员注册一个管理员账户
        :return: 无
        """
        name = self.__get_input("请输入要添加的管理员的姓名", root)
        password = self.__get_input("请输入该账号的密码", root)
        if self.login_level == "Admin":  # 验证用户等级
            wb = load_workbook("data/Admin_names_and_passwords.xlsx")
            sheet = wb.active
            encryptor = FixedIVEncryptor()
            admin_num = len(get_users_and_passwords()["Admin"])
            sheet[f"{get_column_letter(admin_num + 1)}1"] = encryptor.encrypt(name)
            sheet[f"{get_column_letter(admin_num + 1)}2"] = encryptor.encrypt(password)
            wb.save("data/Admin_names_and_passwords.xlsx")

    def add_grader(self, name, password):
        """
        用于管理员或老师注册一个打分员账户
        :param name: 注册的姓名
        :param password: 注册密码
        :return: 无
        """
        if self.login_level == "Admin" or self.login_level == "Teacher":  # 验证用户等级
            wb = load_workbook("data/Grader_names_and_passwords.xlsx")
            sheet = wb.active
            encryptor = FixedIVEncryptor()
            admin_num = len(get_users_and_passwords()["Grader"])
            sheet[f"{get_column_letter(admin_num + 1)}1"] = encryptor.encrypt(name)
            sheet[f"{get_column_letter(admin_num + 1)}2"] = encryptor.encrypt(password)
            wb.save("data/Grader_names_and_passwords.xlsx")

    def add_teacher(self, name, password):
        """
                用于管理员注册一个管理员账户
                :param name: 注册的姓名
                :param password: 注册密码
                :return: 无
                """
        if self.login_level == "Admin":  # 验证用户等级
            wb = load_workbook("data/Teacher_names_and_passwords.xlsx")
            sheet = wb.active
            encryptor = FixedIVEncryptor()
            admin_num = len(get_users_and_passwords()["Teacher"])
            sheet[f"{get_column_letter(admin_num + 1)}1"] = encryptor.encrypt(name)
            sheet[f"{get_column_letter(admin_num + 1)}2"] = encryptor.encrypt(password)
            wb.save("data/Teacher_names_and_passwords.xlsx")

    def delete_admin(self, name: str):
        if self.login_level != "Admin":
            print("❌ 权限不足")
            return

        try:
            # 1. 获取原始的加密字典（密文用户名 -> 密文密码）
            encrypted_dict = get_users_and_passwords()["Admin"]  # 假设这就是 {enc_user: enc_pwd}

            # 2. 创建解密映射：{明文用户名: (密文用户名, 密文密码)}
            encryptor = FixedIVEncryptor()
            plaintext_to_encrypted = {}

            for enc_user, enc_pwd in encrypted_dict.items():
                try:
                    plain_user = encryptor.decrypt(enc_user)
                    # plain_pwd = encryptor.decrypt(enc_pwd)  # 暂时不需要密码
                    plaintext_to_encrypted[plain_user] = enc_user  # 只需密文用户名用于定位
                except IndexError as e:
                    print(e)
                    continue  # 跳过无效条目
            # 3. 如果目标用户不在其中
            if name not in plaintext_to_encrypted:
                print(f"⚠️ 用户 '{name}' 不存在")
                return

            # 4. 找到对应的密文用户名
            target_enc_user = plaintext_to_encrypted[name]

            # 5. 加载 Excel 文件，查找哪一列的 A1、B1... 等于 target_enc_user
            wb = load_workbook("data/Admin_names_and_passwords.xlsx")
            ws = wb.active

            found_col = None
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                if cell_value == target_enc_user:
                    found_col = col_idx
                    break

            if found_col is None:
                print("❌ 数据不一致：密文用户名未在 Excel 中找到")
                return

            # 6. 清空该列
            ws.cell(row=1, column=found_col).value = None
            ws.cell(row=2, column=found_col).value = None

            wb.save("data/Admin_names_and_passwords.xlsx")
            print(f"✅ 成功删除管理员: {name}")

        except Exception as e:
            print(f"❌ 删除失败: {e}")
            import traceback
            traceback.print_exc()

    def delete_grader(self, name: str):
        if self.login_level != "Grader":
            print("❌ 权限不足")
            return

        try:
            # 1. 获取原始的加密字典（密文用户名 -> 密文密码）
            encrypted_dict = get_users_and_passwords()["Grader"]
            encryptor = FixedIVEncryptor()
            plaintext_to_encrypted = {}

            for enc_user, enc_pwd in encrypted_dict.items():
                try:
                    plain_user = encryptor.decrypt(enc_user)
                    # plain_pwd = encryptor.decrypt(enc_pwd)  # 暂时不需要密码
                    plaintext_to_encrypted[plain_user] = enc_user  # 只需密文用户名用于定位
                except IndexError as e:
                    print(e)
                    continue  # 跳过无效条目
            # 3. 如果目标用户不在其中
            if name not in plaintext_to_encrypted:
                print(f"⚠️ 用户 '{name}' 不存在")
                return

            # 4. 找到对应的密文用户名
            target_enc_user = plaintext_to_encrypted[name]

            # 5. 加载 Excel 文件，查找哪一列的 A1、B1... 等于 target_enc_user
            wb = load_workbook("data/Grader_names_and_passwords.xlsx")
            ws = wb.active

            found_col = None
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                if cell_value == target_enc_user:
                    found_col = col_idx
                    break

            if found_col is None:
                print("❌ 数据不一致：密文用户名未在 Excel 中找到")
                return

            # 6. 清空该列
            ws.cell(row=1, column=found_col).value = None
            ws.cell(row=2, column=found_col).value = None

            wb.save("data/Grader_names_and_passwords.xlsx")
            print(f"✅ 成功删除打分员: {name}")

        except Exception as e:
            print(f"❌ 删除失败: {e}")
            import traceback
            traceback.print_exc()

    def delete_teacher(self, name: str):
        if self.login_level != "Teacher":
            print("❌ 权限不足")
            return

        try:
            # 1. 获取原始的加密字典（密文用户名 -> 密文密码）
            encrypted_dict = get_users_and_passwords()["Teacher"]  # 假设这就是 {enc_user: enc_pwd}

            # 2. 创建解密映射：{明文用户名: (密文用户名, 密文密码)}
            encryptor = FixedIVEncryptor()
            plaintext_to_encrypted = {}

            for enc_user, enc_pwd in encrypted_dict.items():
                try:
                    plain_user = encryptor.decrypt(enc_user)
                    # plain_pwd = encryptor.decrypt(enc_pwd)  # 暂时不需要密码
                    plaintext_to_encrypted[plain_user] = enc_user  # 只需密文用户名用于定位
                except IndexError as e:
                    print(e)
                    continue  # 跳过无效条目
            # 3. 如果目标用户不在其中
            if name not in plaintext_to_encrypted:
                print(f"⚠️ 用户 '{name}' 不存在")
                return

            # 4. 找到对应的密文用户名
            target_enc_user = plaintext_to_encrypted[name]

            # 5. 加载 Excel 文件，查找哪一列的 A1、B1... 等于 target_enc_user
            wb = load_workbook("data/Teacher_names_and_passwords.xlsx")
            ws = wb.active

            found_col = None
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col_idx).value
                if cell_value == target_enc_user:
                    found_col = col_idx
                    break

            if found_col is None:
                print("❌ 数据不一致：密文用户名未在 Excel 中找到")
                return

            # 6. 清空该列
            ws.cell(row=1, column=found_col).value = None
            ws.cell(row=2, column=found_col).value = None

            wb.save("data/Teacher_names_and_passwords.xlsx")
            print(f"✅ 成功删除教师: {name}")

        except Exception as e:
            print(f"❌ 删除失败: {e}")
            import traceback
            traceback.print_exc()

    # 接下来要实现 Grader 对数据库的操作方法，首先要用到的属性有login_level, 以及SQL的Leaf类。
    def punch_in(self):
        sql = MyEasySQLite("punch_in.db")
        table = "punch_in"
        sql.create_table(table, column_names=['name', 'punch_days', 'punch_state'])
        if sql.get_data(table, {"name": self.login_name}):
            """
                :param "name": 名字
                :param "punch_days": 打卡的天数

            """
            sql.update_data(table_name=table,
                            new_data={'punch_days': int(sql.get_data(table_name=table,
                                                                     conditions={'name': self.login_name})),
                                      'punch_state': str(not self.__punch_state)},
                            conditions={"name": self.login_name})
        else:  # 直接注册该用户到punch SQL
            sql.update_data(table_name=table,
                            new_data={"punch_days": 1},
                            conditions={"name": self.login_name})
            sql.update_data(table, {"punch_state": "True"},
                            conditions={"name": self.login_name})

    def get_punch_state(self, user_name) -> bool:
        sql = MyEasySQLite("punch_in.db")
        table = "punch_in"
        self.__punch_state = bool(sql.get_data(table_name=table,
                                               conditions={'name': user_name}))
        return self.__punch_state

    def sql_add_data(self, root):
        if self.login_level != "has no login" and self.login_level != "Teacher":
            leaf: Leaf = Leaf(self.login_name)
            for i in leaf.classes:
                for j in range(5):
                    u: str = self.__get_input(f"请输入你要给{i}班添加第{j}天的分数值", root)
                    leaf.add_data(u)
            print("保存中......")
            self.__punch_state = True
            leaf.save()
            print("保存完毕")

    def sql_del_data(self, root):
        """
        用于操作sql的删除操作
        """

        if self.login_level != "has no login" and self.login_level != "Teacher":
            c = self.__get_input("你要删除哪个班的分数", root)
            d = self.__get_input("你要删除这个班第几天的分数", root)
            sql = MyEasySQLite(f"Leaf/{self.login_name}/{get_time()['y']}-{get_time()['m']}.db")
            con = {"class": c, "day": d}
            sql.delete_data(f"{get_time()['d']}", con)

    def sql_update_data(self, root):
        if self.login_level != "has no login" and self.login_level != "Teacher":
            sql = MyEasySQLite(f"Leaf/{self.login_name}/{get_time()['y']}-{get_time()['m']}.db")
            c = self.__get_input("请输入你要替换的班级", root)
            d = self.__get_input("请输入你要替换哪一天的分数", root)
            data = {"score": self.__get_input("请输入你要替换的分数", root)}
            con = {"class": c, "day": d}
            sql.update_data(f"{get_time()['d']}", new_data=data, conditions=con)

    def sql_get_table(self) -> list[str]:
        result = []
        if self.login_level != "has no login" and self.login_level != "Teacher":
            sql = MyEasySQLite(f"Leaf/{self.login_name}/{get_time()['y']}-{get_time()['m']}.db")
            t = sql.get_data(f"{get_time()['d']}")
            tmp: dict[str, str] = t[0]
            for k, v in tmp.items():
                if k == "score":
                    result.append(v)
        return result

    ####################test#############
    @staticmethod
    def get_user_names(__level: str) -> list[str]:
        """
        获取所有账号的明文姓名列表。
        :return: 姓名列表，如 ["张三", "李四"]
        """
        encryptor = FixedIVEncryptor()
        all_users = get_users_and_passwords()
        admin_encrypted_dict = all_users.get(f"{__level}", {})

        admin_names = []
        for encrypted_name in admin_encrypted_dict.keys():
            try:
                plain_name = encryptor.decrypt(encrypted_name)
                admin_names.append(plain_name)
            except ValueError as e:
                # 可选：记录日志或跳过损坏条目
                print(e)
                continue
        return admin_names